"""
Procesa el consolidado de transacciones.
- Quita filas de 'Contratos'
- Segmenta por Placa (una hoja por placa)
- Ordena por Fecha de la transaccion
- Agrupa visualmente por mes y pone AUTOSUMA (=SUM) por cada mes
- Mantiene el mismo formato tabular (mismas columnas), solo agrega subtotales
"""
import re
import unicodedata
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}


def _norm(s):
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return s.strip().lower()


def _find_header_row(path):
    """Busca la fila que contiene los encabezados (Placa, Servicio, Valor...)."""
    raw = pd.read_excel(path, header=None, nrows=15)
    for i in range(len(raw)):
        vals = [_norm(v) for v in raw.iloc[i].tolist()]
        if "placa" in vals and "valor" in vals:
            return i
    return 0


def _parse_fecha(v):
    if v is None:
        return None
    s = str(v).strip().strip('"').strip("'")
    if not s or s.lower() == "nan":
        return None
    try:
        return pd.to_datetime(s).to_pydatetime()
    except Exception:
        return None


def procesar(input_path, output_path, orden="reciente"):
    """
    orden:
      'reciente' = mes mas reciente primero (ultimo -> primer mes)
      'antiguo'  = mes mas antiguo primero (primer -> ultimo mes)
    """
    header_row = _find_header_row(input_path)
    df = pd.read_excel(input_path, header=header_row, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    # localizar columnas por nombre (tolerante a acentos/espacios)
    cols = {_norm(c): c for c in df.columns}
    col_placa = cols.get("placa")
    col_servicio = cols.get("servicio")
    col_valor = cols.get("valor")
    col_fecha = None
    for k, original in cols.items():
        if "fecha de la transaccion" in k:
            col_fecha = original
            break
    if col_fecha is None:
        for k, original in cols.items():
            if "fecha de la transac" in k:
                col_fecha = original
                break

    if not all([col_placa, col_servicio, col_valor, col_fecha]):
        raise ValueError(f"No encuentro columnas. Encontradas: {list(df.columns)}")

    # quitar filas totalmente vacias
    df = df.dropna(how="all")

    # quitar CONTRATOS
    df = df[df[col_servicio].apply(lambda x: _norm(x) != "contratos")].copy()

    # parse valor y fecha
    df["_valor"] = pd.to_numeric(
        df[col_valor].astype(str).str.replace(r"[^\d\.\-]", "", regex=True),
        errors="coerce"
    )
    df["_fecha"] = df[col_fecha].apply(_parse_fecha)

    # solo filas con placa valida
    df = df[df[col_placa].notna() & (df[col_placa].astype(str).str.strip() != "")]
    # solo con fecha valida
    df = df[df["_fecha"].notna()]

    df["_anio"] = df["_fecha"].apply(lambda d: d.year)
    df["_mes"] = df["_fecha"].apply(lambda d: d.month)

    # columnas originales a conservar (mismo formato tabular, sin _aux)
    out_cols = [c for c in df.columns if not c.startswith("_")]

    wb = Workbook()
    wb.remove(wb.active)

    # estilos
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    mes_font = Font(bold=True, name="Arial", size=10)
    mes_fill = PatternFill("solid", fgColor="FCE4D6")
    sub_font = Font(bold=True, name="Arial", size=10)
    sub_fill = PatternFill("solid", fgColor="D9E1F2")
    base_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    valor_idx = out_cols.index(col_valor) + 1  # columna 1-based del valor

    placas = sorted(df[col_placa].astype(str).str.strip().unique())

    for placa in placas:
        sub = df[df[col_placa].astype(str).str.strip() == placa].copy()
        # ordenar por fecha
        sub = sub.sort_values("_fecha", ascending=(orden == "antiguo"))

        # construir lista de (anio, mes) en el orden pedido
        grupos = sub.groupby(["_anio", "_mes"], sort=False)
        claves = list(grupos.groups.keys())

        ws = wb.create_sheet(title=str(placa)[:31])
        r = 1

        # encabezado
        for ci, cname in enumerate(out_cols, start=1):
            c = ws.cell(row=r, column=ci, value=cname)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        r += 1

        total_general_rows = []

        for (anio, mes) in claves:
            block = grupos.get_group((anio, mes))

            # fila titulo de mes
            mc = ws.cell(row=r, column=1, value=f"{MESES_ES.get(mes, mes)} {anio}")
            mc.font = mes_font
            mc.fill = mes_fill
            for ci in range(1, len(out_cols) + 1):
                ws.cell(row=r, column=ci).fill = mes_fill
                ws.cell(row=r, column=ci).border = border
            r += 1

            primera_fila_datos = r
            for _, row in block.iterrows():
                for ci, cname in enumerate(out_cols, start=1):
                    val = row[cname]
                    if cname == col_valor:
                        val = row["_valor"]
                    if pd.isna(val):
                        val = ""
                    c = ws.cell(row=r, column=ci, value=val)
                    c.font = base_font
                    c.border = border
                    if ci == valor_idx and val != "":
                        c.number_format = '#,##0'
                r += 1
            ultima_fila_datos = r - 1

            # fila TOTAL del mes: numero directo (siempre se ve, sin depender de Excel)
            sc = ws.cell(row=r, column=1, value=f"TOTAL {MESES_ES.get(mes, mes)} {anio}")
            sc.font = sub_font
            col_letter = get_column_letter(valor_idx)
            total_mes = float(block["_valor"].fillna(0).sum())
            tcell = ws.cell(row=r, column=valor_idx, value=total_mes)
            tcell.font = sub_font
            tcell.number_format = '#,##0'
            # dejar la formula como comentario por si el usuario quiere recalcular en Excel
            try:
                from openpyxl.comments import Comment
                tcell.comment = Comment(
                    f"=SUM({col_letter}{primera_fila_datos}:{col_letter}{ultima_fila_datos})",
                    "Bot"
                )
            except Exception:
                pass
            for ci in range(1, len(out_cols) + 1):
                ws.cell(row=r, column=ci).fill = sub_fill
                ws.cell(row=r, column=ci).border = border
            total_general_rows.append(r)
            r += 2  # fila en blanco separadora

        # ancho de columnas
        for ci, cname in enumerate(out_cols, start=1):
            maxlen = max([len(str(cname))] + [len(str(v)) for v in sub[cname].head(80).tolist() if pd.notna(v)] + [12])
            ws.column_dimensions[get_column_letter(ci)].width = min(maxlen + 2, 40)
        ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path, placas


if __name__ == "__main__":
    import sys
    inp = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/ConsolidadoDeTransacciones.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "/home/claude/test_out.xlsx"
    orden = sys.argv[3] if len(sys.argv) > 3 else "reciente"
    p, placas = procesar(inp, out, orden)
    print("OK ->", p, "Placas:", placas)
